import numpy as np
import pickle
from keras.models import load_model
from usol.deep import usolDlib
from keras_vggface.vggface import VGGFace
from keras.preprocessing.image import ImageDataGenerator, load_img
import os
import random
import openpyxl
import pandas as pd
from sklearn.utils import shuffle
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from PIL import Image
from sklearn.model_selection import KFold, train_test_split
from sklearn.model_selection import cross_val_score


test_dir = '../dataset/pre_over_100/people/testConfidence'

usolDlib = usolDlib()
myModel = load_model('hs_model.h5')

resnet_vgg = VGGFace(model='resnet50', include_top=False,
                    input_shape=(224, 224, 3), pooling='max', weights='vggface')
resnet_vgg.summary()
datagen = ImageDataGenerator()
batch_size = 100
labelCount = None

def makeFeature(generator, n):
    _features = np.zeros(shape=(n, 2048))
    _labels = np.zeros(shape=(n, labelCount))

    i = 0
    for inputs_batch, labels_batch in generator:
        features_batch = resnet_vgg.predict(inputs_batch)
        features_batch = (features_batch - features_batch.min()) / features_batch.max() - features_batch.min()
        _features[i * batch_size: (i + 1) * batch_size] = features_batch
        _labels[i * batch_size: (i + 1) * batch_size] = labels_batch
        print('extract feature -------------------> ' + str((i + 1) * batch_size))
        i += 1
        if i * batch_size >= n:
            break

    # _features = np.reshape(_features, (n, 1*1*2048))
    return _features, _labels

test_generator = datagen.flow_from_directory(
    test_dir,
    target_size=(224, 224),
    batch_size=batch_size,
    class_mode='categorical',
    shuffle=False)

day = '0604'
labelCount = test_generator.class_indices.__len__()
if os.path.isfile("confiXY" + day + ".npz"):
    print("-----------------load feature!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!-----------------")
    data= np.load('confiXY' + day + '.npz')
    features, labels = data['x'], data['y']
else:
    print("-----------------save feature!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!-----------------")
    features, labels = makeFeature(test_generator, test_generator.classes.size)
    np.savez('confiXY' + day, x=features, y=labels)

data = np.load('dataXY.npz')
compare_features, compare_labels = data['x'], data['y']

fnames = test_generator.filenames
predictions = myModel.predict_classes(features)
prob = myModel.predict(features)

INDEX_FILE_NAME = 'hs_model_label.pkl'
with open(INDEX_FILE_NAME, 'rb') as f:
    idx2label = pickle.load(f)

wb = openpyxl.Workbook()
ws = wb.active

#df = pd.DataFrame(data={'prob' : np.array([], dtype=float), 'ox' : np.array([], dtype=int), 'l2min' : np.array([], dtype=float)})
df = pd.DataFrame(columns=("prob","l2min","ox"))
for i in range(len(predictions)):
    pred_class = np.argmax(prob[i])
    pred_label = idx2label[pred_class]

    #one hot encoding된 트레이닝 라벨값 인트로 변경하여 l2계산을 위해 비교 할 피쳐index 가져오기
    idx = np.where(np.argmax(compare_labels, axis=1) == pred_class)[0]
    #트레이닝 feature중 랜덤 10개 선정
    ridx = [random.choice(idx) for j in range(10)]

    l2dists = []
    for k in ridx:
        dist = np.linalg.norm(features[i].reshape(-1) - compare_features[k], axis=None, ord=None)
        l2dists.append(dist)
    #print("L2 distance(max " + str(np.max(l2dists)) + " / min " + str(np.min(l2dists)) + ")")

    print('Original label:{}, Prediction :{}, prob : {:.3f}, L2 max : {:.5f}, L2 min : {:.5f}, L2 avg : {:.5f}'.format(
        fnames[i].split('\\')[0],
        pred_label,
        prob[i][pred_class],
        np.max(l2dists),
        np.min(l2dists),
        np.average(l2dists)))
    name = fnames[i].split('\\')[0]
    if name == pred_label:
        ox = 1
    else:
        ox = 0

    df.loc[i] = [prob[i][pred_class], np.min(l2dists), ox]
    '''
    name = fnames[i].split('\\')[0]    
    ws['a' + str(i +2)] = name
    ws['b' + str(i + 2)] = pred_label
    ws['c' + str(i + 2)] = prob[i][pred_class]
    ws['d' + str(i + 2)] = np.max(l2dists)
    ws['e' + str(i + 2)] = np.min(l2dists)
    ws['f' + str(i + 2)] = np.average(l2dists)    
    '''
#wb.save('Confidence1.xlsx')

#df = shuffle(df)
train_set, test_set = train_test_split(df, test_size=0.25, random_state=93)
train_data =  train_set.drop('ox', axis=1)
train_data =  train_data.drop('l2min', axis=1)
train_label = train_set['ox']
test_data =  test_set.drop('ox', axis=1)
test_data =  test_data.drop('l2min', axis=1)
test_label = test_set['ox']

#n_jobs  = 병렬화 옵션, cv =교차검증, random_state 랜덤 값 지속성 보장
#k_fold = 교차검증 옵션
clf  =  KNeighborsClassifier (n_neighbors = 5)
k_fold = KFold(n_splits=10, shuffle=True, random_state=5)
scoring = 'accuracy'
score = cross_val_score(clf, train_data, train_label, cv=k_fold, n_jobs=1, scoring=scoring)
print(score)
print(round(np.mean(score)*100, 2))


clf = DecisionTreeClassifier()
scoring = 'accuracy'
score = cross_val_score(clf, train_data, train_label, cv=k_fold, n_jobs=1, scoring=scoring)
print(score)
print(round(np.mean(score)*100, 2))

clf = RandomForestClassifier(n_estimators=5)
scoring = 'accuracy'
score = cross_val_score(clf, train_data, train_label, cv=k_fold, n_jobs=1, scoring=scoring)
print(score)
print(round(np.mean(score)*100, 2))

clf = SVC()
scoring = 'accuracy'
score = cross_val_score(clf, train_data, train_label, cv=k_fold, n_jobs=1, scoring=scoring)
print(score)
print(round(np.mean(score)*100,2))


clf = DecisionTreeClassifier()
clf.fit(train_data, train_label)
print(clf.score(train_data, train_label))
print(clf.score(test_data, test_label))

with open('hs_svm_work_unknown.pkl', 'wb') as f:
    pickle.dump(clf, f)
